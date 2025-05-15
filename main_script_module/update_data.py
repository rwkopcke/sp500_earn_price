'''This program reads selected data from S&P, sp-500-eps-est.xlsx
        https://www.spglobal.com/spdji/en/search/?query=index+earnings&activeTab=all
   and from the 10-year TIPS rate from FRED: 
        https://fred.stlouisfed.org/series/DFII10
   It writes these data as polars dataframes to .parquet files
        and writes a record of the files that it has read and writen
        as a dictionary to a .json file
        
   see sp_paths.py
        for the addresses of the files within this project are declared
        for advice in resetting paths for debugging and for reinitializing
        the project's output files

   The polars dataframes in input_output
        the latest projections of earnings for the
        S&P500 within each quarter since late 2017. 
   A separate polars dataframe contains
        the actual earnings and the value of the index for each quarter 
        beginning in 1988. This dataframe also contains actual values for 
        operating margins, revenues, book values, dividends, and other 
        actual data reported by S&P, plus actual values for the 10-year TIPS.
   
   The addresses of documents for this project appear in this program's 
   project directory: S&P500_PE/sp500_pe/sp_paths.py
'''

#######################  Parameters  ###################################
import sys
import gc

import polars as pl
import polars.selectors as cs
import json
from openpyxl import load_workbook

from main_script_module import sp_paths as sp
from main_script_module import update_record
from helper_func_module import helper_func as hp
from helper_func_module import read_data_func as rd

from dataclasses import dataclass

@dataclass(frozen= True)
class Fixed_Project_Parameters:
    ARCHIVE_RR_FILE = False

    # data from "ESTIMATES&PEs" wksht
    RR_COL_NAME = 'real_int_rate'
    YR_QTR_NAME = 'yr_qtr'
    PREFIX_OUTPUT_FILE_NAME = 'sp-500-eps-est'
    EXT_OUTPUT_FILE_NAME = '.parquet'

    SHT_EST_NAME = "ESTIMATES&PEs"
    COLUMN_NAMES = ['date', 'price', 'op_eps', 'rep_eps',
                    'op_p/e', 'rep_p/e', '12m_op_eps', '12m_rep_eps']
    PROJ_COLUMN_NAMES = ['date', 'op_eps', 'rep_eps',
                        'op_p/e', 'rep_p/e', '12m_op_eps', '12m_rep_eps']

    SHT_QTR_NAME = "QUARTERLY DATA"
    COLUMN_NAMES_QTR = ['date', 'div_ps', 'sales_ps',
                        'bk_val_ps', 'capex_ps', 'divisor']

    SHT_IND_NAME = 'SECTOR EPS'

    # NB ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # all search (row or col) "keys" should be None or lists
    # all column indexes in skip lists below are zero-based ('A' is 0)
    # all specific individual column designations are letters
    # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    ACTUAL_KEYS = ['ACTUALS', 'Actuals']
    
    SHT_EST_DATE_PARAMS = {
        'date_keys' : ['Date', 'Data as of the close of:'],
        'value_col_1' : 'D',
        'date_key_2' : ACTUAL_KEYS,
        'value_col_2' : 'B',
        'column_names' : COLUMN_NAMES,
        'yr_qtr_name' : YR_QTR_NAME
    }

    SHT_HIST_PARAMS = {
        'act_key' : ACTUAL_KEYS,
        'end_key' : None,
        'first_col' : 'A',
        'last_col' : 'J',
        'skip_cols' : [4, 7],
        'column_names' : COLUMN_NAMES,
        'yr_qtr_name' : YR_QTR_NAME
    }

    MARG_KEY = 'QTR'
    SHT_BC_MARG_PARAMS = {
        'row_key': MARG_KEY,
        'first_col': 'A',
        'stop_col_key': None,
        'stop_row_data_offset': 4,
        'yr_qtr_name': YR_QTR_NAME
    }

    SHT_IND_PARAMS = {
        'first_row_op': 6,
        'first_row_rep': 63,
        'num_inds': 12,
        'start_col_key': None,
        'stop_col_key': None
    }

    SHT_QTR_PARAMS = {
        'act_key' : ['END'],
        'end_key' : None,
        'first_col' : 'A',
        'last_col' : 'I',
        'skip_cols' : [1, 2, 7],
        'column_names' : COLUMN_NAMES_QTR,
        'yr_qtr_name' : YR_QTR_NAME
    }

    SHT_EST_PROJ_DATE_PARAMS = {
        'date_keys' : ['Date', 'Data as of the close of:'],
        'value_col_1' : 'D', 
        'date_key_2' : None, 
        'value_col_2' : None,
        'column_names' : None,
        'yr_qtr_name' : YR_QTR_NAME
    }

    SHT_EST_PROJ_PARAMS = {
        'act_key' : ['ESTIMATES'],
        'end_key' : ACTUAL_KEYS,
        'first_col' : 'A',
        'last_col' : 'J',
        'skip_cols' : [1, 4, 7],
        'column_names' : PROJ_COLUMN_NAMES,
        'yr_qtr_name' : YR_QTR_NAME
    }

    SHT_FRED_PARAMS = {
        'first_row': 12,
        'col_1': 'A',
        'col_2': 'B',
        'yr_qtr_name': YR_QTR_NAME,
        'rr_col_name': RR_COL_NAME
    }


#######################  MAIN Function  ###############################

def update():
    '''create or update earnings, p/e, real int rates, margins, etc.
       from 'sp-500-eps-est ...' files
    '''
    # set immutable constants
    env = Fixed_Project_Parameters()
    
    # load file containing record_dict: record of files seen previously
    #   if record_dict does not exist, create an empty dict to initialize
    record_dict, files_to_read_list = update_record()

## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++              
## +++++  fetch the historical data  +++++++++++++++++++++++++++++++++++++++
## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++ 

    print('\n================================================')
    print(f'Updating historical data from: {record_dict["latest_used_file"]}')
    print(f'in directory: \n{sp.path.INPUT_DIR}')
    print('================================================\n')
    
## ACTUAL DATA from existing .parquet file (to be updated below)
    # the rows (qtrs) not to be updated are the rows that
    # do not contain null in the op_eps col
    # put the yr_qtr for these rows in the set rows_no_update
    if sp.path.OUTPUT_HIST_ADDR.exists():
        actual_df = pl.read_parquet(sp.path.OUTPUT_HIST_ADDR)
    
        rows_not_to_update = set(pl.Series(actual_df
                                    .drop_nulls(subset='op_eps')
                                    .select(pl.col(env.YR_QTR_NAME)))
                                .to_list())
    else:
        rows_not_to_update = []
    
## REAL INTEREST RATES, eoq, from FRED DFII10
    active_workbook = load_workbook(filename= sp.path.INPUT_RR_ADDR,
                                    read_only= True,
                                    data_only= True)
    active_sheet = active_workbook.active
    real_rt_df = rd.fred_reader(active_sheet,
                                **env.SHT_FRED_PARAMS)

## NEW HISTORICAL DATA
    ## WKSHT with new historical values for P and E from new excel file
    latest_file_addr = sp.path.INPUT_DIR / record_dict["latest_used_file"]
    active_workbook = load_workbook(filename= latest_file_addr,
                                    read_only= True,
                                    data_only= True)
    # most recent date and prices
    active_sheet = active_workbook[env.SHT_EST_NAME]

    # new_df dates and latest prices, beyond historical data
    name_date, add_df = rd.read_sp_date(active_sheet, 
                                        **env.SHT_EST_DATE_PARAMS,
                                        include_prices= True)
    # load new historical data
    # omit rows whose yr_qtr appears in the rows_no_update list
    df = rd.sp_loader(active_sheet,
                      rows_not_to_update,
                      **env.SHT_HIST_PARAMS)
    
    # if any date is None, halt
    if (name_date is None or
        any([item is None
            for item in add_df['date']])):
        
        print('\n============================================')
        print(f'{latest_file_addr} \nmissing history date')
        print(f'Name_date: {name_date}')
        print(actual_df['date'])
        print('============================================\n')
        sys.exit()
        
    # update add_df with new historical data
    add_df = pl.concat([add_df, df], how= "diagonal")
               
    # build new_df with rr (to update rows in actual_df)
    # merge new real_rates with new p and e
    add_df = add_df.join(real_rt_df, 
                         how="left", 
                         on=[env.YR_QTR_NAME],
                         coalesce= True)
    del real_rt_df
    del df
    gc.collect()
        
## MARGINS add to new df
    margins_df = rd.margin_loader(active_sheet,
                                  rows_not_to_update,
                                  **env.SHT_BC_MARG_PARAMS)
    
    add_df = add_df.join(margins_df, 
                         how="left", 
                         on= env.YR_QTR_NAME,
                         coalesce= True)
    del margins_df
    gc.collect()

## QUARTERLY DATA add to new_df
    active_sheet = active_workbook[env.SHT_QTR_NAME]

    # ensure all dtypes (if not string or date-like) are float32
    # some dtype are null when all col entries in short df are null
    qtrly_df = rd.sp_loader(active_sheet,
                            rows_not_to_update,
                            **env.SHT_QTR_PARAMS)\
                 .cast({~(cs.temporal() | cs.string()): pl.Float32,
                        cs.datetime(): pl.Date})
    
    add_df = add_df.join(qtrly_df,  
                         how= "left", 
                         on= [env.YR_QTR_NAME],
                         coalesce= True)
    
    del qtrly_df
    gc.collect()
    
## ACTUAL_DF update: remove rows to be updated and concat with add_df
    # align cols of actual_df with add_df
    # ensure rows do not overlap
    if sp.path.OUTPUT_HIST_ADDR.exists():
        actual_df = pl.concat([add_df.filter(
                                     ~pl.col(env.YR_QTR_NAME)
                                      .is_in(rows_not_to_update))
                                  .sort(by= env.YR_QTR_NAME),
                               actual_df.select(
                                      add_df.columns)
                                  .filter(pl.col(env.YR_QTR_NAME)
                                      .is_in(rows_not_to_update))
                                  .sort(by= env.YR_QTR_NAME)],
                               how= 'vertical')
    else:
        actual_df = add_df.sort(by= env.YR_QTR_NAME)
    
    del add_df
    gc.collect()

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
## INDUSTRIAL DATA
    # read stored data
    if sp.path.OUTPUT_IND_ADDR.exists():
        ind_df = pl.read_parquet(sp.path.OUTPUT_IND_ADDR)\
                        .sort(by= 'year', descending= True)
                    
        years_no_update = set(pl.Series(ind_df
                                .drop_nulls(subset='SP500_rep_eps')
                                .select(pl.col('year')))
                                .to_list())
    else:
        years_no_update = []
    
    # find new industry data
    active_sheet = active_workbook[env.SHT_IND_NAME]
    add_ind_df = rd.industry_loader(active_sheet,
                                    years_no_update,
                                    **env.SHT_IND_PARAMS)
    # add col with Q4 value of real_int_rate each year from actual_df
    add_ind_df = \
        add_ind_df.join(
                actual_df.select([env.YR_QTR_NAME, 'real_int_rate'])
                            .filter(pl.col(env.YR_QTR_NAME)
                            .map_elements(lambda x: x[-1:]=='4',
                                            return_dtype= bool))
                            .with_columns(pl.col(env.YR_QTR_NAME)
                            .map_elements(lambda x: x[0:4],
                                            return_dtype= str)
                            .alias('year'))
                            .drop(env.YR_QTR_NAME),
                        on= 'year',
                        how= 'left',
                        coalesce= True)\
                 .sort(by= 'year', descending= True)\
                 .cast({~cs.string() : pl.Float32})
    
    if sp.path.OUTPUT_IND_ADDR.exists():
        years = pl.Series(add_ind_df['year']).to_list()
        ind_df = pl.concat([add_ind_df,
                            ind_df.filter(~pl.col('year')
                                            .is_in(years))],
                            how= 'vertical')
    else:
        ind_df = add_ind_df.sort(by= 'year', descending= True)

## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
## +++++ update projection files +++++++++++++++++++++++++++++++++++++++++++
## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    # ordinarily a very short list
    # loop through files_to_read, fetch projections of earnings for each date
    add_proj_yr_qtrs_list = []
    failure_to_read_lst = []
    for file in files_to_read_list:
        # echo file name and address to console
        active_workbook = load_workbook(filename= sp.path.INPUT_DIR / file,
                                        read_only= True,
                                        data_only= True)
        active_sheet = active_workbook[env.SHT_EST_NAME]
        print(f'\n input file: {file}')    
        
# projections of earnings
        # read date of projection, no prices or other data
        name_date, _ = \
            rd.read_sp_date(active_sheet,
                            **env.SHT_EST_PROJ_DATE_PARAMS)
        name_date = name_date.date()
    
        # load projections for the date
        proj_df = rd.sp_loader(active_sheet,
                               [],
                               **env.SHT_EST_PROJ_PARAMS)

        # if any date is None, abort and continue
        if (name_date is None or
            any([item is None
                for item in proj_df['date']])):
            print('\n============================================')
            print('In main(), projections:')
            print(f'Skipped sp-500 {name_date} missing projection date')
            print('============================================\n')
            failure_to_read_lst.append(file)
            continue
        
        year_quarter = \
            f'{name_date.year}-Q{(int(name_date.month) - 1) // 3 + 1}'
            
        add_proj_yr_qtrs_list.append(year_quarter)
        

## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
## +++++ write files +++++++++++++++++++++++++++++++++++++++++++++++++++++++
## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

## +++++  write proj_df  ++++++++++++++++++++++++++++++++++++++++++++++++++
        output_file_name = \
            f'{env.PREFIX_OUTPUT_FILE_NAME} {name_date}{env.EXT_OUTPUT_FILE_NAME}'
        record_dict['output_proj_files'].append(output_file_name)
        output_file_address = sp.path.OUTPUT_PROJ_DIR / output_file_name
        
        print(year_quarter)
        print(f'output file: {output_file_name}')
        print(proj_df['yr_qtr'].to_list())
        print()
        
        with output_file_address.open('w') as f:
            proj_df.write_parquet(f)
            
## +++++ write history file ++++++++++++++++++++++++++++++++++++++++++++
    # move any existing hist file in output_dir to backup
    if sp.path.OUTPUT_HIST_ADDR.exists():
        sp.path.OUTPUT_HIST_ADDR.rename(sp.path.BACKUP_HIST_ADDR)
        print('\n============================================')
        print(f'Moved history file from: \n{sp.path.OUTPUT_HIST_ADDR}')
        print(f'to: \n{sp.path.BACKUP_HIST_ADDR}')
        print('============================================\n')
    else:
        print('\n============================================')
        print(f'Found no history file at: \n{sp.path.OUTPUT_HIST_ADDR}')
        print(f'Wrote no history file to: \n{sp.path.BACKUP_HIST_ADDR}')
        print('============================================\n')
        
    # write actual_df, the historical data, into the output file
    with sp.path.OUTPUT_HIST_ADDR.open('w') as f:
        actual_df.write_parquet(f)
    print('\n============================================')
    print(f'Wrote history file to: \n{sp.path.OUTPUT_HIST_ADDR}')
    print('============================================\n')
    
## +++++ write industry file ++++++++++++++++++++++++++++++++++++++++++++
    # move any existing industry file in output_dir to backup
    if sp.path.OUTPUT_IND_ADDR.exists():
        sp.path.OUTPUT_IND_ADDR.rename(sp.path.BACKUP_IND_ADDR)
        print('\n============================================')
        print(f'Moved industry file from: \n{sp.path.OUTPUT_IND_ADDR}')
        print(f'to: \n{sp.path.BACKUP_IND_ADDR}')
        print('============================================\n')
    else:
        print('\n============================================')
        print(f'Found no industry file at: \n{sp.path.OUTPUT_IND_ADDR}')
        print(f'Wrote no industry file to: \n{sp.path.BACKUP_IND_ADDR}')
        print('============================================\n')
        
    # write ind_df, the industry data, into the output file
    with sp.path.OUTPUT_IND_ADDR.open('w') as f:
        ind_df.write_parquet(f)
    print('\n============================================')
    print(f'Wrote industry file to: \n{sp.path.OUTPUT_IND_ADDR}')
    print('============================================\n')
            
## +++++ update archive ++++++++++++++++++++++++++++++++++++++++
    # archive all input files -- uses Path() variables
    # https://sysadminsage.com/python-move-file-to-another-directory/
    print('\n============================================')
    for file in new_files_list:
        input_address = sp.path.INPUT_DIR / file
        if input_address.exists():
            input_address.rename(sp.path.ARCHIVE_DIR / file)
            print(f"Archived: {input_address}")
            
        else:
            print(f"\nWARNING")
            print(f"Tried: {input_address}")
            print(f'Address does not exist\n')
    print('============================================\n')
    
    if env.ARCHIVE_RR_FILE:
        sp.path.INPUT_RR_ADDR.rename(sp.path.ARCHIVE_DIR / sp.path.INPUT_RR_FILE)
        print('\n============================================')
        print(f"Archived: \n{sp.path.INPUT_RR_FILE}")
        print('============================================\n')
    
            
    # list should begin with most recent items
    # more efficient search for items to edit above
    record_dict['prev_files'].sort(reverse= True)
    record_dict['prev_used_files'].sort(reverse= True)
    record_dict['output_proj_files'].sort(reverse= True)
    record_dict['proj_yr_qtrs'].sort(reverse= True)
            
## store record_dict
    with sp.path.RECORD_DICT_ADDR.open('w') as f:
        json.dump(record_dict, f)
    print('\n====================================================')
    print('Saved record_dict to file')
    print(f'{sp.path.RECORD_DICT_ADDR}')
    print(f'\nlatest_used_file: {record_dict['latest_used_file']}\n')
    print(f'output_proj_files: \n{record_dict['output_proj_files'][:6]}\n')
    print(f'prev_used_files: \n{record_dict['prev_used_files'][:6]}\n')
    print(f'prev_files: \n{record_dict['prev_files'][:6]}\n')
    print(f'proj_yr_qtrs: \n{record_dict['proj_yr_qtrs'][:6]}\n')
    print('====================================================\n')
 
    print('\n====================================================')
    print('Retrieval is complete\n')
    
    n = len(files_to_read_list)
    m = len(failure_to_read_lst)
    print(f'{n} new input files read and saved')
    print(f'from {sp.path.INPUT_DIR}')
    print(f'  to {sp.path.OUTPUT_DIR}\n')
    print(f'{m} files not read and saved:\n')
    print(failure_to_read_lst)
    print('====================================================')
    
    return